from tkinter import *
import tkinter.font as tkFont
import pandas as pd
import openpyxl
import pathlib


class Excel_File_Modification:
    def __init__(self, root):

        self.ModificationFrame = root

        self.fontStyle = tkFont.Font(family="Comic Sans MS", size=11, weight = "bold")


        self.ModificationButton = Button(self.ModificationFrame, text = "Modify Excel File", padx = 203, pady = 10, command = self.Modify_Excel_Function, font = self.fontStyle)
        self.ModificationButton.grid(row = 0, column = 0, padx = 370, pady = 10)
        #self.ModificationButton.grid_propagate(False)

    def Modify_Excel_Function(self):

        self.ModificationButton.destroy()
        #Add a tag
        def Add_Tag_Function():
            Text_Tag_To_Be_Entered = Add_Tag_Entry.get()

            filePath = pathlib.Path('Balance Sheet.xlsx')
            Existing_Tags_DataFrame = pd.read_excel(filePath, sheet_name = 'Tags', header = None)

            Add_Tag_Entry.delete(0, "end")


            if [Text_Tag_To_Be_Entered] in Existing_Tags_DataFrame.iloc[:,[0]].values.tolist():

                return

            else:

                #print(Existing_Tags_DataFrame.shape)
                Existing_Tags_DataFrame_plus_New_Tag = pd.concat([Existing_Tags_DataFrame, pd.DataFrame([[Text_Tag_To_Be_Entered]])], ignore_index = True)
                # df = pd.concat([df,df_new_line], ignore_index=True)
                # creating excel writer for storing the calculation outputs into different excel sheets
                #print(Existing_Tags_DataFrame_plus_New_Tag)
                writer = pd.ExcelWriter('Balance Sheet.xlsx', mode = 'a', if_sheet_exists = 'replace', engine = "openpyxl")

                # writing the transactions within user input dates into a new excel sheet
                Existing_Tags_DataFrame_plus_New_Tag.to_excel(writer, sheet_name = "Tags", index = None, startrow = 0, startcol = 0, header = None)

                writer.close()#save()

            return
        Add_Tag_Button = Button(self.ModificationFrame, text = "Add This Tag > ", bg = 'white', width = 25, command = Add_Tag_Function, font = self.fontStyle)
        Add_Tag_Button.grid(row = 0, column = 0, sticky = W, padx = 17)
        Add_Tag_Entry = Entry(self.ModificationFrame, bg = 'white', width = 25, bd = 5, font = self.fontStyle)
        Add_Tag_Entry.grid(row = 0, column = 1, sticky = W, padx = 11)
        #Delete a tag
        def Delete_Tag_Function():
            Text_Tag_To_Be_Deleted = Delete_Tag_Entry.get()

            filePath = pathlib.Path('Balance Sheet.xlsx')
            Existing_Tags_DataFrame = pd.read_excel(filePath, sheet_name = 'Tags', header = None)

            Delete_Tag_Entry.delete(0, "end")


            if [Text_Tag_To_Be_Deleted] in Existing_Tags_DataFrame.iloc[:,[0]].values.tolist():

                print(Existing_Tags_DataFrame)

                Final_Tags_After_Deletion = Existing_Tags_DataFrame.loc[Existing_Tags_DataFrame.iloc[:,[0]].values != [Text_Tag_To_Be_Deleted]]

                print(Final_Tags_After_Deletion)

                writer = pd.ExcelWriter('Balance Sheet.xlsx', mode = 'a', if_sheet_exists = 'replace', engine = "openpyxl")

                # writing the transactions within user input dates into a new excel sheet
                Final_Tags_After_Deletion.to_excel(writer, sheet_name = "Tags", index = None, startrow = 0, startcol = 0, header = None)

                writer.close()#save()

            else:
                return


            return
        Delete_Tag_Button = Button(self.ModificationFrame, text = "Delete This Tag > ", bg = 'white', width = 25, command = Delete_Tag_Function, font = self.fontStyle)
        Delete_Tag_Button.grid(row = 1, column = 0, sticky = W, padx = 17)
        Delete_Tag_Entry = Entry(self.ModificationFrame, bg = 'white', width = 25, bd = 5, font = self.fontStyle)
        Delete_Tag_Entry.grid(row = 1, column = 1, sticky = W, padx = 10)
        #Add a bank Account
        def Add_Account_Function():

            filePath = pathlib.Path('Balance Sheet.xlsx')
            Existing_DataFrame = pd.read_excel(filePath, sheet_name = 'Sheet', header = 0)

            Existing_DataFrame.insert (Existing_DataFrame.shape[1] - 2, Add_Account_Name_Entry.get(), int(Add_Account_Balance_Entry.get()))


            Add_Account_Name_Entry.delete(0, "end")
            Add_Account_Balance_Entry.delete(0, "end")


            writer = pd.ExcelWriter('Balance Sheet.xlsx', mode = 'a', if_sheet_exists = 'replace', engine = "openpyxl")

            Existing_DataFrame['Date'] = pd.to_datetime(Existing_DataFrame['Date'])
            n = 1
            for x in Existing_DataFrame.iloc[1:len(Existing_DataFrame), 0]:
                Existing_DataFrame.iloc[ n , 0 ] = x.strftime("%d-%b-%Y")
                n += 1
            Existing_DataFrame['Date'] = Existing_DataFrame['Date'].astype(str)

            # writing the transactions within user input dates into a new excel sheet
            Existing_DataFrame.to_excel(writer, sheet_name = "Sheet", index = None, startrow = 0, startcol = 0)

            writer.close()#save()


            return
        Add_Account_Button = Button(self.ModificationFrame, text = "< Add This \nAccount", bg = 'white', pady = 9, width = 16, command = Add_Account_Function, font = self.fontStyle)
        Add_Account_Button.grid(row = 0, column = 4, sticky = W, rowspan = 2, padx = 10)
        Add_Account_Name_Entry = Entry(self.ModificationFrame, bg = 'white', width = 25, bd = 5, font = self.fontStyle)
        Add_Account_Name_Entry.grid(row = 0, column = 3, sticky = W, padx = 10)
        Add_Account_Balance_Entry = Entry(self.ModificationFrame, bg = 'white', width = 25, bd = 5, font = self.fontStyle)
        Add_Account_Balance_Entry.grid(row = 1, column = 3, sticky = W, padx = 10)
        Add_Account_Name_Text = Label(self.ModificationFrame, text = "Account Name: ", bg = 'white', width = 25, font = self.fontStyle)
        Add_Account_Name_Text.grid(row = 0, column = 2, sticky = E, padx = 10)
        Add_Account_Balance_Text = Label(self.ModificationFrame, text = "Account Balance: ", bg = 'white', width = 25, font = self.fontStyle)
        Add_Account_Balance_Text.grid(row = 1, column = 2, sticky = E, padx = 10)

        #delete last excel entry row
        def Delete_Last_Excel_Entry_Function():
            # enter your file path
            filePath = pathlib.Path('Balance Sheet.xlsx')

            # load excel file
            book = openpyxl.load_workbook(filePath)

            # select the sheet
            sheet = book['Sheet']

            # sheet.max_row is the maximum number
            # of rows that the sheet have
            # delete_row() method removes rows, first parameter represents row
            # number and sencond parameter represents number of rows
            # to delete from the row number
            sheet.delete_rows(sheet.max_row, 1)

            book.save(filePath)

            return
        Delete_Last_Excel_Entry_Button = Button(self.ModificationFrame, text = "Delete Last \nTransaction Entry", bg = 'white', pady = 9, width = 16, command = Delete_Last_Excel_Entry_Function, font = self.fontStyle)
        Delete_Last_Excel_Entry_Button.grid(row = 0, column = 5, sticky = W, padx = 39, rowspan = 2)
        return



    def start(self):
        self.root.mainloop()
